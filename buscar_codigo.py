import openpyxl
from pathlib import Path


codigo_buscado = input("¿Qué código buscas?: ").upper()
carpeta = Path("D:\\")

for archivo in carpeta.glob("*.xlsx"):
    try:
        wb = openpyxl.load_workbook(str(archivo), data_only=True)
        
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            
            for fila_num, fila in enumerate(ws.iter_rows(min_row=6, max_col=5), start=6):
                if fila[0].value and str(fila[0].value).upper() == codigo_buscado:
                    print(f"✓ ENCONTRADO: {archivo.name}")
                    print(f"  Hoja: {nombre_hoja}, Fila: {fila_num}")
                    print(f"  Código: {fila[0].value}")
                    print(f"  Denominación: {fila[1].value}")
                    print()
    except Exception as e:
        print(f"⚠ Error en {archivo.name}: {e}")

print("Búsqueda completada")