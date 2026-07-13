import os
import django
import re
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import Titulos, Centros, MateriasAvaliadas

def traducir_a_gallego(nombre):
    traducciones = {
        'Ambientales': 'Ambientais',
        'Ingeniería': 'Enxeñaría',
        'Enfermería': 'Enfermaría',
        'Tecnología': 'Tecnoloxía',
        'Arqueología': 'Arqueoloxía',
        'Ciudades': 'Cidades',
        'Inteligencia': 'Intelixencia',
        'Bachillerato': 'Bacharelato',
        'Biotecnología': 'Biotecnoloxía',
        'Biología': 'Bioloxía',
        'Abordaje': 'Abordaxe',
        'Genética': 'Xenética',
        'Energía': 'Enerxía',
        'Industriales': 'Industriáis',
        'Industrial': 'Industrial',
        'Extranjera': 'Extranxeira',
        'Realidad': 'Realidade',
        'Geoespacial': 'Xeoespacial',
        'Gestión': 'Xestión',
        'Desarrollo': 'Desenvolvemento',
        'Sostenible': 'Sostible',
        'Nanotecnología': 'Nanotecnoloxía',
        'Derecho': 'Dereito',
        'Diseño': 'Deseño',
        'Extranjeras': 'Extranxeiras',
        'Lenguas': 'Linguas',
        'Traducción': 'Tradución',
        'Filología': 'Filoloxía',
        'Lingüística': 'Lingüística',
        'Lenguas': 'Linguas',
        'Extranjeras': 'Estranxeiras',
        'Traducción': 'Tradución',
        'Interpretación': 'Interpretación',
        'Lingüística': 'Lingüística',
        'Literatura': 'Literatura',
        'Dramática': 'Dramática',
        'Escénicas': 'Escénicas',
        'Enseñanza': 'Ensino',
        'Español': 'Español',
        'Segunda': 'Segunda',
        'Aplicaciones': 'Aplicacións',
        'Relaciones': 'Relacións',
        'Internacionales': 'Internacionais',
        'PCEO': 'PCEO',
        'Máster': 'Máster',
        'Grado': 'Grao',
        'Automática': 'Automática',
        'Mecánica': 'Mecánica', 
        'Biomédica': 'Biomédica',
        'Graduado': 'Grao'
    }
    nombre_traducido = nombre.lower()
    for castellano, gallego in traducciones.items():
        nombre_traducido = nombre_traducido.replace(castellano.lower(), gallego.lower())
    return nombre_traducido

def buscar_titulo(nombre, codigo_centro):
    if not nombre:
        return None
    nombre_traducido = traducir_a_gallego(nombre)
    palabras_genericas = {'en', 'de', 'y', 'e', 'grado', 'grao', 'máster', 'universitario', 'para', 'por', 'con'}
    palabras_clave = [p for p in nombre_traducido.lower().split()
                      if len(p) > 3 and p not in palabras_genericas]
    centro = Centros.objects.filter(codigo=codigo_centro).first()
    titulos = Titulos.objects.filter(centro=centro) if centro else Titulos.objects.all()
    mejor_match = None
    mejor_coincidencias = 0
    for titulo in titulos:
        denominacion = titulo.denominacion.lower()
        coincidencias = sum(1 for palabra in palabras_clave if palabra in denominacion)
        if coincidencias > mejor_coincidencias:
            mejor_coincidencias = coincidencias
            mejor_match = titulo
    return mejor_match if mejor_coincidencias >= 1 else None

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques

def auditar_arquivo(archivo_path, f):
    match = re.match(r'^(\d{3})', archivo_path.name)
    codigo_centro = match.group(1) if match else None

    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)

    if 'Anexos 1' not in wb.sheetnames:
        f.write("⚠ Non se atopou hoja 'Anexos 1'\n\n")
        return 0, 0

    ws = wb['Anexos 1']
    bloques = detectar_bloques(ws)

    n_ok = 0
    n_faltantes = 0

    for col_inicio in bloques:
        nombre_titulo = ws.cell(row=6, column=col_inicio).value
        titulo = buscar_titulo(nombre_titulo, codigo_centro) if nombre_titulo else None

        if not titulo:
            f.write(f"✗ Título non encontrado: {nombre_titulo}\n")
            n_faltantes += 1
            continue

        for fila_num in range(6, ws.max_row + 1):
            codigo_asig = ws.cell(row=fila_num, column=col_inicio + 1).value
            nome_asig = ws.cell(row=fila_num, column=col_inicio + 2).value

            if not codigo_asig and not nome_asig:
                break

            if not codigo_asig:
                continue

            materia = MateriasAvaliadas.objects.filter(
                codigo=str(codigo_asig).strip()
            ).first()

            if materia:
                n_ok += 1
            else:
                f.write(f"✗ Materia non importada: {titulo} —  {codigo_asig} — {nome_asig}\n")
                n_faltantes += 1

    return n_ok, n_faltantes

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])

    salida = Path('auditoria_materias.txt')
    total_ok = 0
    total_faltantes = 0

    with open(salida, 'w', encoding='utf-8') as f:
        f.write("AUDITORÍA DE MATERIAS\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            n_ok, n_faltantes = auditar_arquivo(archivo_path, f)
            f.write(f"RESUMEN: {n_ok} correctas, {n_faltantes} faltantes\n\n")
            total_ok += n_ok
            total_faltantes += n_faltantes

        f.write("=" * 100 + "\n")
        f.write(f"TOTAL: {total_ok} correctas, {total_faltantes} faltantes\n")

    print(f"Resultado en: {salida}")

if __name__ == '__main__':
    main()