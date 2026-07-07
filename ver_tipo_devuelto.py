import openpyxl
from pathlib import Path

archivo = Path("D:\\107 DE-02 Cadro de mando 25-26 Aeronautica Modificado.xlsx")
wb = openpyxl.load_workbook(str(archivo), data_only=True)
ws = wb['Centro']

# Leer fila 7 (primer indicador)
# fila = list(ws.iter_rows(min_row=20, max_row=20, values_only=True))[0]
print("Fila 7 completa:")
fila = list(ws.iter_rows(min_row=20, max_row=20, values_only=True))[0]
for col, valor in enumerate(fila, 0):
    print(f"  Col {col}: {valor}")

print(f"Col 5 (meta): {fila[5]} - Tipo: {type(fila[4])}")
print(f"Col 6 (resultado): {fila[6]} - Tipo: {type(fila[5])}")
print("\nEncabezado fila 5:")
for col, valor in enumerate([cell.value for cell in ws[5]], 1):
    if valor:
        print(f"  Col {col}: {valor}")