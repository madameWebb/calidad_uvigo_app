# comparar_excels.py (en la raíz del proyecto)

import openpyxl
from pathlib import Path

def main():
    archivo_origen = input("Excel a limpiar: ").strip()
    archivo_referencia = input("Excel de referencia: ").strip()
    
    origen = Path(archivo_origen)
    referencia = Path(archivo_referencia)
    
    if not origen.exists() or not referencia.exists():
        print("Archivos no encontrados")
        return

    # Leer archivos
    wb_origen = openpyxl.load_workbook(str(origen), data_only=True)
    ws_origen = wb_origen.active
    
    wb_ref = openpyxl.load_workbook(str(referencia), data_only=True)
    ws_ref = wb_ref.active

    # Extraer plan_sigma de referencia
    plan_sigma_ref = set()
    for fila in ws_ref.iter_rows(min_row=2, values_only=True):
        plan_sigma = fila[0]
        if plan_sigma:
            plan_sigma_ref.add(str(plan_sigma).strip())

    # Encabezado
    encabezado = [cell.value for cell in ws_origen[1]]
    
    # Filas nuevas
    filas_nuevas = []
    for fila in ws_origen.iter_rows(min_row=2, values_only=True):
        plan_sigma = fila[0]
        if plan_sigma and str(plan_sigma).strip() not in plan_sigma_ref:
            filas_nuevas.append(fila)

    # Crear nuevo Excel
    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    
    for col_num, valor in enumerate(encabezado, 1):
        ws_nuevo.cell(row=1, column=col_num, value=valor)
    
    for row_num, fila in enumerate(filas_nuevas, 2):
        for col_num, valor in enumerate(fila, 1):
            ws_nuevo.cell(row=row_num, column=col_num, value=valor)

    archivo_nuevo = Path(f'{origen.stem}_limpio.xlsx')
    wb_nuevo.save(str(archivo_nuevo))

    print(f'De {ws_origen.max_row - 1} filas, {len(filas_nuevas)} son nuevas.')
    print(f'Archivo: {archivo_nuevo}')

if __name__ == '__main__':
    main()