import os
import django
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import Indicadores, Codigos, Titulos

ANEXOS_CURSOS = {
    'Anexos 1': '23/24',
    'Anexos 2': '24/25',
    'Anexos 3': '25/26',
}

def detectar_anexos(wb):
    cursos = {}
    anexos_encontrados = sorted([h for h in wb.sheetnames if h.lower().startswith('anexo')])
    cursos_orden = ['23/24', '24/25', '25/26']
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

def procesar_arquivo(archivo_path, f):
    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    cursos = detectar_anexos(wb)
    
    if not cursos:
        f.write("⚠ Non se atoparon hojas de Anexos\n\n")
        return

    for nombre_hoja, orixe_datos in cursos.items():
        ws = wb[nombre_hoja]
        f.write(f"HOJA: {nombre_hoja} → {orixe_datos}\n")
        f.write("-" * 100 + "\n")

        codigo_indicador = ws.cell(row=2, column=2).value
        if not codigo_indicador:
            f.write("⚠ Non hai código de indicador na fila 2\n\n")
            continue

        indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
        f.write(f"Indicador: {codigo_indicador} → {'✓ ' + indicador.denominacion if indicador else '✗ Non encontrado'}\n\n")

        for fila_num in range(6, ws.max_row + 1):
            xescampus = ws.cell(row=fila_num, column=2).value
            nombre_titulo = ws.cell(row=fila_num, column=3).value

            if not xescampus and not nombre_titulo:
                break

            codigo_obj = Codigos.objects.filter(xescampus=str(xescampus).strip()).first() if xescampus else None

            if not codigo_obj:
                if nombre_titulo:
                    titulo = Titulos.objects.filter(denominacion__icontains=nombre_titulo).first()
                    if titulo:
                        f.write(f"  Fila {fila_num}: {xescampus} → ✗ xescampus non encontrado\n")
                        f.write(f"    Título por nome: ✓ {titulo.denominacion}\n")
                    else:
                        f.write(f"  Fila {fila_num}: {xescampus} → ✗ xescampus non encontrado\n")
                        f.write(f"    Título por nome: ✗ '{nombre_titulo}' non encontrado\n")
                else:
                    f.write(f"  Fila {fila_num}: sen xescampus nin nome de título\n")
            else:
                titulo = codigo_obj.titulo
                f.write(f"  Fila {fila_num}: {xescampus}\n")
                f.write(f"    Título: ✓ {titulo}\n")

            excelentes = ws.cell(row=fila_num, column=4).value or 0
            notables = ws.cell(row=fila_num, column=5).value or 0
            favorables = ws.cell(row=fila_num, column=6).value or 0
            desfavorables = ws.cell(row=fila_num, column=7).value or 0
            totales = ws.cell(row=fila_num, column=8).value or 0

            f.write(f"    Excelentes: {excelentes}, Notables: {notables}, Favorables: {favorables}, Desfavorables: {desfavorables}, Totais: {totales}\n\n")

        f.write("\n")

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)
    
    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])
    
    if not archivos:
        print("Non hai arquivos .xlsx na carpeta")
        return

    salida = Path('importar_pdis_preview.txt')
    with open(salida, 'w', encoding='utf-8') as f:
        f.write("PREVIEW PDIs — CARPETA COMPLETA\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            procesar_arquivo(archivo_path, f)

    print(f"Resultado en: {salida}")

if __name__ == '__main__':
    main()