import os
import django
import re
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import Titulos, Centros

def traducir_a_gallego(nombre):
    traducciones = {
        'Ambientales': 'Ambientais',
        'Ingeniería': 'Enxeñaría',
        'Enfermería': 'Enfermaría',
        'Tecnología': 'Tecnoloxía',
        'Arqueología': 'Arqueoloxía',
        'Ciudades': 'Cidades',
        'Inteligencia': 'Intelixencia',
        'Bachillerato': 'Bacharelato',
        'Biotecnología': 'Biotecnoloxía',
        'Biología': 'Bioloxía',
        'Abordaje': 'Abordaxe',
        'Genética': 'Xenética',
        'Energía': 'Enerxía',
        'Industriales': 'Industriáis',
        'Industrial': 'Industrial',
        'Extranjera': 'Extranxeira',
        'Realidad': 'Realidade',
        'Geoespacial': 'Xeoespacial',
        'Gestión': 'Xestión',
        'Desarrollo': 'Desenvolvemento',
        'Sostenible': 'Sostible',
        'Nanotecnología': 'Nanotecnoloxía',
        'Derecho': 'Dereito',
        'Diseño': 'Deseño',
        'Extranjeras': 'Extranxeiras',
        'Lenguas': 'Linguas',
        'Traducción': 'Tradución',
        'Filología': 'Filoloxía',
        'Lingüística': 'Lingüística',
        'Lenguas': 'Linguas',
        'Extranjeras': 'Estranxeiras',
        'Traducción': 'Tradución',
        'Interpretación': 'Interpretación',
        'Lingüística': 'Lingüística',
        'Literatura': 'Literatura',
        'Dramática': 'Dramática',
        'Escénicas': 'Escénicas',
        'Enseñanza': 'Ensino',
        'Español': 'Español',
        'Segunda': 'Segunda',
        'Aplicaciones': 'Aplicacións',
        'Relaciones': 'Relacións',
        'Internacionales': 'Internacionais',
        'PCEO': 'PCEO',
        'Máster': 'Máster',
        'Grado': 'Grao',
        'Automática': 'Automática',
        'Mecánica': 'Mecánica', 
        'Biomédica': 'Biomédica',
        'Graduado': 'Grao',
        'Electrónica': 'Electrónica', 
        'Mecatrónica': 'Mecatrónica', 
        'Sustentabilidade': 'Sustentabilidade'

    }

    nombre_traducido = nombre.lower()
    for castellano, gallego in traducciones.items():
        nombre_traducido = nombre_traducido.replace(castellano.lower(), gallego.lower())
    return nombre_traducido

def buscar_titulo(nombre, codigo_centro):
    if not nombre:
        return None
    
    nombre_traducido = traducir_a_gallego(nombre)
    palabras_genericas = {'en', 'de', 'y', 'e', 'universitario', 'para', 'o'}
    palabras_clave = [p for p in nombre_traducido.lower().split()
                      if len(p) > 3 and p not in palabras_genericas]

    centro = Centros.objects.filter(codigo=codigo_centro).first()

    if centro:
        titulos = Titulos.objects.filter(centro=centro)
    else:
        titulos = Titulos.objects.all()

    mejor_match = None
    mejor_coincidencias = 0

    for titulo in titulos:
        denominacion = titulo.denominacion.lower()
        coincidencias = sum(1 for palabra in palabras_clave if palabra in denominacion)

        if coincidencias > mejor_coincidencias:
            mejor_coincidencias = coincidencias
            mejor_match = titulo

    if mejor_coincidencias >= 3:
        return mejor_match

    return None

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques

def procesar_arquivo(archivo_path, f):
    match = re.match(r'^(\d{3})', archivo_path.name)
    codigo_centro = match.group(1) if match else None

    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)

    if 'Anexos 1' not in wb.sheetnames:
        f.write("⚠ Non se atopou hoja 'Anexos 1'\n\n")
        return

    ws = wb['Anexos 1']
    bloques = detectar_bloques(ws)

    f.write(f"Bloques detectados: {bloques}\n\n")

    for col_inicio in bloques:
        nombre_titulo = ws.cell(row=6, column=col_inicio).value
        
        # Saltar PCEO
        if nombre_titulo and str(nombre_titulo).upper().startswith('PCEO'):
            print(f'Saltado PCEO: {nombre_titulo}')
            f.write(f"⚠ Saltado PCEO: {nombre_titulo}\n")
            continue
    
        titulo = buscar_titulo(nombre_titulo, codigo_centro) if nombre_titulo else None
        
        f.write(f"BLOQUE col {col_inicio}: {nombre_titulo}\n")
        f.write(f"  Título: {'✓ ' + str(titulo) if titulo else '✗ Non encontrado'}\n")
        f.write("-" * 80 + "\n")

        for fila_num in range(6, ws.max_row + 1):
            codigo_asig = ws.cell(row=fila_num, column=col_inicio + 1).value
            nome_asig = ws.cell(row=fila_num, column=col_inicio + 2).value

            if not codigo_asig and not nome_asig:
                break

            f.write(f"  Fila {fila_num}: {codigo_asig} — {nome_asig}\n")

        f.write("\n")

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])

    if not archivos:
        print("Non hai arquivos .xlsx")
        return

    salida = Path('importar_materias_preview.txt')
    with open(salida, 'w', encoding='utf-8') as f:
        f.write("PREVIEW MATERIAS — CARPETA COMPLETA\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            procesar_arquivo(archivo_path, f)


    print(f"Resultado en: {salida}")

if __name__ == '__main__':
    main()