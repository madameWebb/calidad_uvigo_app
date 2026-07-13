import os
import django
import re
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import MateriasAvaliadas, Indicadores, SeguementoMaterias

ANEXOS_ESTANDAR = {'Anexos 1', 'Anexos 2', 'Anexos 3'}

def detectar_anexos(wb):
    cursos = {}
    cursos_orden = ['23/24', '24/25', '25/26']
    anexos_encontrados = sorted([h for h in wb.sheetnames if h.strip() in ANEXOS_ESTANDAR])
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques

def auditar_arquivo(archivo_path, f):
    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    cursos = detectar_anexos(wb)

    if not cursos:
        f.write("⚠ Non se atoparon Anexos estándar\n\n")
        return 0, 0

    n_ok = 0
    n_faltantes = 0

    for nombre_hoja, orixe_datos in cursos.items():
        ws = wb[nombre_hoja]
        bloques = detectar_bloques(ws)
        f.write(f"HOJA: {nombre_hoja} → {orixe_datos}\n")
        f.write("-" * 100 + "\n")

        for col_inicio in bloques:
            nombre_titulo = ws.cell(row=6, column=col_inicio).value

            if nombre_titulo and str(nombre_titulo).upper().startswith('PCEO'):
                continue

            codigos_raw = ws.cell(row=2, column=col_inicio).value
            if not codigos_raw:
                continue

            codigos = [c.strip() for c in str(codigos_raw).split('\n') if c.strip()]
            indicadores = [Indicadores.objects.filter(codigo=c).first() for c in codigos]
            indicadores = [i for i in indicadores if i]

            for fila_num in range(6, ws.max_row + 1):
                codigo_materia = ws.cell(row=fila_num, column=col_inicio + 1).value
                if not codigo_materia:
                    break

                materia = MateriasAvaliadas.objects.filter(codigo=str(codigo_materia).strip()).first()
                if not materia:
                    f.write(f"✗ Materia {codigo_materia} non encontrada\n")
                    n_faltantes += 1
                    continue

                for indicador in indicadores:
                    seg = SeguementoMaterias.objects.filter(
                        materia=materia,
                        indicador=indicador,
                        orixe_datos=orixe_datos
                    ).first()

                    if seg:
                        n_ok += 1
                    else:
                        f.write(f"✗ Fila {fila_num}: {codigo_materia} — {indicador.codigo} ({orixe_datos}) FALTA\n")
                        n_faltantes += 1

        f.write("\n")

    return n_ok, n_faltantes

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])

    salida = Path('auditoria_seguimentos_materias.txt')
    total_ok = 0
    total_faltantes = 0

    with open(salida, 'w', encoding='utf-8') as f:
        f.write("AUDITORÍA SEGUIMENTOS MATERIAS\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            n_ok, n_faltantes = auditar_arquivo(archivo_path, f)
            f.write(f"RESUMEN: {n_ok} correctos, {n_faltantes} faltantes\n\n")
            total_ok += n_ok
            total_faltantes += n_faltantes

        f.write("=" * 100 + "\n")
        f.write(f"TOTAL: {total_ok} correctos, {total_faltantes} faltantes\n")

    print(f"Resultado en: {salida}")
    print(f"Total correctos: {total_ok}")
    print(f"Total faltantes: {total_faltantes}")

if __name__ == '__main__':
    main()