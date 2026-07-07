import os
import django
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
import re
from gestion.models import Seguimentos, SeguimentosTitulos, Centros, Indicadores

def detectar_cursos(ws):
    cursos = {}
    encabezado = [cell.value for cell in ws[4]]
    
    for col_num, valor in enumerate(encabezado, 1):
        if not valor:
            continue
        valor_str = str(valor).strip()
        
        if '23/24' in valor_str or '2023/2024' in valor_str or '2023-24' in valor_str or '23-24' in valor_str or 'Curso X' == valor_str:
            cursos['23/24'] = (col_num, col_num + 1, col_num + 2)
        elif '24/25' in valor_str or '2024/2025' in valor_str or '2024-25' in valor_str or '24-25' in valor_str or 'Curso X+1' in valor_str:
            cursos['24/25'] = (col_num, col_num + 1, col_num + 2)
        elif '25/26' in valor_str or '2025/2026' in valor_str or '2025-26' in valor_str or '25-26' in valor_str or 'Curso X+2' in valor_str:
            cursos['25/26'] = (col_num, col_num + 1, col_num + 2)
    
    return cursos

SHEETS_EXCLUIDAS = {'Portada', 'Anexos 1', 'Anexos 2', 'Anexos 3', 'Resumo', 'Procedementos'}

def auditar_archivo(archivo_path, f):
    match = re.match(r'^(\d{3})', archivo_path.name)
    if not match:
        f.write(f"⚠ No puedo extraer código de {archivo_path.name}\n\n")
        return 0, 0

    codigo_centro = match.group(1)
    centro = Centros.objects.filter(codigo=codigo_centro).first()
    if not centro:
        f.write(f"⚠ Centro {codigo_centro} no existe\n\n")
        return 0, 0

    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    

    n_ok = 0
    n_faltantes = 0

    f.write(f"ARQUIVO: {archivo_path.name}\n")
    f.write(f"CENTRO: {centro.denominacion}\n")
    f.write("=" * 100 + "\n\n")

    # --- Hoja Centro ---
    if 'Centro' in wb.sheetnames:
        ws = wb['Centro']
        cursos = detectar_cursos(ws)
        if not cursos:
            f.write("⚠ No se detectaron cursos en hoja Centro\n\n")
        else:
            f.write("HOJA: Centro\n")
            f.write("-" * 100 + "\n")
                
        for fila_num, fila in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
            codigo_indicador = fila[0]
            
            if not codigo_indicador or 'Indicadores' in str(codigo_indicador):
                continue
            
            indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
            if not indicador:
                f.write(f"? Fila {fila_num}: Indicador {codigo_indicador} no existe en BD\n")
                continue

            for orixe_datos in cursos.keys():
                seg = Seguimentos.objects.filter(
                    centro=centro,
                    indicador=indicador,
                    orixe_datos=orixe_datos
                ).first()

                if seg:
                    n_ok += 1
                else:
                    f.write(f"✗ Fila {fila_num}: {codigo_indicador} ({orixe_datos}) FALTA\n")
                    n_faltantes += 1

    # --- Hojas de títulos ---
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja in SHEETS_EXCLUIDAS or nombre_hoja == 'Centro':
            continue

        titulo = None
        for t in centro.titulos.all():
            ws = wb[nombre_hoja]
            for fila in ws.iter_rows(max_row=5, values_only=True):
                for celda in fila:
                    if celda and isinstance(celda, str) and t.denominacion in celda:
                        titulo = t
                        break
                if titulo:
                    break
            if titulo:
                break

        if titulo is None:
            continue

        ws = wb[nombre_hoja]
        cursos = detectar_cursos(ws)
        if not cursos:
            f.write(f"⚠ No se detectaron cursos en hoja {nombre_hoja}\n")
            continue
        f.write(f"\nHOJA: {nombre_hoja} — {titulo.denominacion}\n")
        f.write("-" * 100 + "\n")

        for fila_num, fila in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
            codigo_indicador = fila[0]
            
            if not codigo_indicador or 'Indicadores' in str(codigo_indicador):
                continue
            
            indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
            if not indicador:
                f.write(f"? Fila {fila_num}: Indicador {codigo_indicador} no existe en BD\n")
                continue

            for orixe_datos in cursos.keys():
                seg = SeguimentosTitulos.objects.filter(
                    titulo=titulo,
                    indicador=indicador,
                    orixe_datos=orixe_datos
                ).first()

                if seg:
                    n_ok += 1
                else:
                    f.write(f"✗ Fila {fila_num}: {codigo_indicador} ({orixe_datos}) FALTA\n")
                    n_faltantes += 1

    f.write(f"\nRESUMEN ARQUIVO: {n_ok} correctos, {n_faltantes} faltantes\n")
    f.write("\n" + "=" * 100 + "\n\n")

    return n_ok, n_faltantes

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)
    
    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])
    
    if not archivos:
        print("Non hai arquivos .xlsx na carpeta")
        return

    salida = Path('auditoria_seguimentos.txt')
    total_ok = 0
    total_faltantes = 0

    with open(salida, 'w', encoding='utf-8') as f:
        f.write("AUDITORÍA GLOBAL DE SEGUIMENTOS\n")
        f.write("=" * 100 + "\n\n")

        for archivo in archivos:
            print(f"Auditando: {archivo.name}")
            n_ok, n_faltantes = auditar_archivo(archivo, f)
            total_ok += n_ok
            total_faltantes += n_faltantes

        f.write("RESUMEN GLOBAL\n")
        f.write("=" * 100 + "\n")
        f.write(f"Total correctos: {total_ok}\n")
        f.write(f"Total faltantes: {total_faltantes}\n")

    print(f"\nResultado en: {salida}")
    print(f"Total correctos: {total_ok}")
    print(f"Total faltantes: {total_faltantes}")

if __name__ == '__main__':
    main()