import os
import django
import re
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import MateriasAvaliadas, Indicadores

ANEXOS_ESTANDAR = {'Anexos 1', 'Anexos 2', 'Anexos 3'}

def detectar_anexos(wb):
    cursos = {}
    cursos_orden = ['23/24', '24/25', '25/26']
    anexos_encontrados = sorted([h for h in wb.sheetnames if h.strip() in ANEXOS_ESTANDAR])
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques

def procesar_arquivo(archivo_path, f):
    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    cursos = detectar_anexos(wb)

    if not cursos:
        f.write("⚠ Non se atoparon Anexos estándar\n\n")
        return

    for nombre_hoja, orixe_datos in cursos.items():
        ws = wb[nombre_hoja]
        bloques = detectar_bloques(ws)
        f.write(f"HOJA: {nombre_hoja} → {orixe_datos}\n")
        f.write("-" * 100 + "\n")

        for col_inicio in bloques:
            nombre_titulo = ws.cell(row=6, column=col_inicio).value

            if nombre_titulo and str(nombre_titulo).upper().startswith('PCEO'):
                f.write(f"⚠ Saltado PCEO: {nombre_titulo}\n")
                continue

            codigos_raw = ws.cell(row=2, column=col_inicio).value
            if not codigos_raw:
                continue

            codigos = [c.strip() for c in str(codigos_raw).split('\n') if c.strip()]
            indicadores = []
            for codigo in codigos:
                ind = Indicadores.objects.filter(codigo=codigo).first()
                if ind:
                    indicadores.append(ind)
                    f.write(f"  ✓ Indicador: {codigo} — {ind.denominacion}\n")
                else:
                    f.write(f"  ✗ Indicador: {codigo} NON ENCONTRADO\n")

            for fila_num in range(6, ws.max_row + 1):
                codigo_materia = ws.cell(row=fila_num, column=col_inicio + 1).value
                if not codigo_materia:
                    break

                materia = MateriasAvaliadas.objects.filter(codigo=str(codigo_materia).strip()).first()
                taxa1 = ws.cell(row=fila_num, column=col_inicio + 4).value
                taxa2 = ws.cell(row=fila_num, column=col_inicio + 5).value
                meta1 = ws.cell(row=fila_num, column=col_inicio + 6).value
                meta2 = ws.cell(row=fila_num, column=col_inicio + 7).value

                if materia:
                    f.write(f"  ✓ Fila {fila_num}: {codigo_materia} — taxa1={taxa1}, taxa2={taxa2}, meta1={meta1}, meta2={meta2}\n")
                else:
                    f.write(f"  ✗ Fila {fila_num}: {codigo_materia} NON ENCONTRADA\n")

        f.write("\n")

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])

    salida = Path('seguimentos_materias_preview.txt')
    with open(salida, 'w', encoding='utf-8') as f:
        f.write("PREVIEW SEGUIMENTOS MATERIAS\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            procesar_arquivo(archivo_path, f)

    print(f"Resultado en: {salida}")

if __name__ == '__main__':
    main()