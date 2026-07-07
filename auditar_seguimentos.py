import os
import django
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
import re
from gestion.models import Seguimentos, Centros, Indicadores

def detectar_cursos(ws):
    cursos = {}
    encabezado = [cell.value for cell in ws[4]]
    
    for col_num, valor in enumerate(encabezado, 1):
        if not valor:
            continue
        valor_str = str(valor).strip()
        
        if '23/24' in valor_str or '2023/2024' in valor_str:
            cursos['23/24'] = (col_num, col_num + 1, col_num + 2)
        elif '24/25' in valor_str or '2024/2025' in valor_str:
            cursos['24/25'] = (col_num, col_num + 1, col_num + 2)
        elif '25/26' in valor_str or '2025/2026' in valor_str:
            cursos['25/26'] = (col_num, col_num + 1, col_num + 2)
        elif 'Curso X+2' in valor_str:
            cursos['25/26'] = (col_num, col_num + 1, col_num + 2)
        elif 'Curso X+1' in valor_str:
            cursos['24/25'] = (col_num, col_num + 1, col_num + 2)
        elif 'Curso X' in valor_str:
            cursos['23/24'] = (col_num, col_num + 1, col_num + 2)
    
    return cursos

def main():
    archivo = input("Ruta del Excel: ").strip()
    archivo_path = Path(archivo)
    
    if not archivo_path.exists():
        print("Archivo no encontrado")
        return

    match = re.match(r'^(\d{3})', archivo_path.name)
    if not match:
        print("No puedo extraer código del archivo")
        return
    
    codigo_centro = match.group(1)
    centro = Centros.objects.filter(codigo=codigo_centro).first()
    if not centro:
        print(f"Centro {codigo_centro} no existe")
        return

    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    ws = wb['Centro']
    cursos = detectar_cursos(ws)

    salida = Path('auditoria_seguimentos.txt')
    with open(salida, 'w', encoding='utf-8') as f:
        f.write(f"AUDITORÍA DE SEGUIMENTOS — {centro.denominacion}\n")
        f.write("=" * 100 + "\n\n")

        n_ok = 0
        n_faltantes = 0

        for fila_num, fila in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
            codigo_indicador = fila[0]
            
            if not codigo_indicador or 'Indicadores' in str(codigo_indicador):
                continue
            
            indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
            if not indicador:
                f.write(f"? Fila {fila_num}: Indicador {codigo_indicador} no existe en BD\n")
                continue

            for orixe_datos in cursos.keys():
                seg = Seguimentos.objects.filter(
                    centro=centro,
                    indicador=indicador,
                    orixe_datos=orixe_datos
                ).first()

                if seg:
                    n_ok += 1
                else:
                    f.write(f"✗ Fila {fila_num}: {codigo_indicador} - {indicador.denominacion} ({orixe_datos}) FALTA\n")
                    n_faltantes += 1

        f.write("\n" + "=" * 100 + "\n")
        f.write(f"RESUMEN: {n_ok} correctos, {n_faltantes} faltantes\n")

    print(f"Resultado en: {salida}")

if __name__ == '__main__':
    main()