import os
import django
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import Indicadores, Codigos, Titulos, AvaliacionsPdis

ANEXOS_ESTANDAR = {'Anexos 1', 'Anexos 2', 'Anexos 3'}

def detectar_anexos(wb):
    cursos = {}
    cursos_orden = ['23/24', '24/25', '25/26']
    anexos_encontrados = sorted([h for h in wb.sheetnames if h.strip() in ANEXOS_ESTANDAR])
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

def auditar_arquivo(archivo_path, f):
    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    cursos = detectar_anexos(wb)

    if not cursos:
        f.write("⚠ Non se atoparon Anexos estándar\n\n")
        return 0, 0

    n_ok = 0
    n_faltantes = 0

    for nombre_hoja, orixe_datos in cursos.items():
        ws = wb[nombre_hoja]
        f.write(f"HOJA: {nombre_hoja} → {orixe_datos}\n")
        f.write("-" * 100 + "\n")

        codigo_indicador = ws.cell(row=2, column=2).value
        if not codigo_indicador:
            f.write("⚠ Non hai código de indicador\n\n")
            continue

        indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
        if not indicador:
            f.write(f"⚠ Indicador {codigo_indicador} non encontrado\n\n")
            continue

        for fila_num in range(6, ws.max_row + 1):
            xescampus = ws.cell(row=fila_num, column=2).value
            nombre_titulo = ws.cell(row=fila_num, column=3).value

            if not xescampus and not nombre_titulo:
                break

            codigo_obj = Codigos.objects.filter(xescampus=str(xescampus).strip()).first() if xescampus else None

            if not codigo_obj:
                if nombre_titulo:
                    titulo_obj = Titulos.objects.filter(denominacion__icontains=nombre_titulo).first()
                else:
                    titulo_obj = None
            else:
                titulo_obj = codigo_obj.titulo

            if not titulo_obj:
                f.write(f"✗ Fila {fila_num}: {xescampus} / '{nombre_titulo}' non encontrado\n")
                n_faltantes += 1
                continue

            avaliacion = AvaliacionsPdis.objects.filter(
                titulo=titulo_obj,
                orixe_datos=orixe_datos
            ).first()

            if avaliacion:
                n_ok += 1
            else:
                f.write(f"✗ Fila {fila_num}: {titulo_obj} ({orixe_datos}) FALTA en BD\n")
                n_faltantes += 1

        f.write(f"\n")

    return n_ok, n_faltantes

def main():
    carpeta = input("Ruta da carpeta: ").strip()
    carpeta_path = Path(carpeta)

    if not carpeta_path.exists():
        print("Carpeta non encontrada")
        return

    archivos = sorted([f for f in carpeta_path.glob('*.xlsx') if not f.name.startswith('~$')])

    if not archivos:
        print("Non hai arquivos .xlsx")
        return

    salida = Path('auditoria_pdis.txt')
    total_ok = 0
    total_faltantes = 0

    with open(salida, 'w', encoding='utf-8') as f:
        f.write("AUDITORÍA DE AVALIACIÓNS PDI\n")
        f.write("=" * 100 + "\n\n")

        for archivo_path in archivos:
            f.write(f"\nARQUIVO: {archivo_path.name}\n")
            f.write("=" * 100 + "\n\n")
            n_ok, n_faltantes = auditar_arquivo(archivo_path, f)
            f.write(f"RESUMEN: {n_ok} correctos, {n_faltantes} faltantes\n\n")
            total_ok += n_ok
            total_faltantes += n_faltantes

        f.write("=" * 100 + "\n")
        f.write(f"TOTAL: {total_ok} correctos, {total_faltantes} faltantes\n")

    print(f"Resultado en: {salida}")
    print(f"Total correctos: {total_ok}")
    print(f"Total faltantes: {total_faltantes}")

if __name__ == '__main__':
    main()