import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gestion.models import Centros, Titulos, Indicadores, Seguimentos, SeguimentosTitulos, Localizadores

User = get_user_model()

SHEETS_EXCLUIDAS = {'Portada', 'Anexos 1', 'Anexos 2', 'Anexos 3', 'Resumo', 'Procedementos'}


def celda_segura(fila, indice):
    if indice < len(fila):
        return fila[indice]
    return None


def limpiar_valor(valor):
    if valor is None:
        return None
    if isinstance(valor, str) and valor.strip() in ('----', '', '-'):
        return None
    return valor


def a_decimal(valor):
    valor = limpiar_valor(valor)
    if valor is None:
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        return None


def detectar_cursos(ws):
    cursos = {}  # ← Siempre devuelve dict, nunca None
    encabezado = [cell.value for cell in ws[4]]
    
    for col_num, valor in enumerate(encabezado, 1):
        if not valor:
            continue
        valor_str = str(valor).strip()
        
        if '23/24' in valor_str or '2023/2024' in valor_str or '2023-24' in valor_str or '23-24' in valor_str or 'Curso X' == valor_str:
            cursos['23/24'] = (col_num, col_num + 1, col_num + 2)
        elif '24/25' in valor_str or '2024/2025' in valor_str or '2024-25' in valor_str or '24-25' in valor_str or 'Curso X+1' in valor_str:
            cursos['24/25'] = (col_num, col_num + 1, col_num + 2)
        elif '25/26' in valor_str or '2025/2026' in valor_str or '2025-26' in valor_str or '25-26' in valor_str or 'Curso X+2' in valor_str:
            cursos['25/26'] = (col_num, col_num + 1, col_num + 2)
    
    return cursos

def valor_seguimento(valor):
    """Convierte el valor de una celda a decimal."""
    if valor is None:
        return None
    valor_str = str(valor).strip()
    if valor_str in ('', '-'):
        return None
    if valor_str in ('----', 'NA', 'na'):
        return Decimal('-1')
    try:
        return Decimal(valor_str)
    except (InvalidOperation, ValueError):
        return None

class Command(BaseCommand):
    help = 'Importa seguimentos desde los Excel'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        
        usuario = User.objects.filter(is_superuser=True).first()
        if usuario is None:
            self.stderr.write('No hay superusuario.')
            return
        
        archivos = list(carpeta.glob('*.xlsx'))
        if not archivos:
            self.stderr.write(f'No hay archivos .xlsx en {carpeta}')
            return

        for ruta in archivos:
            self.procesar_archivo(ruta, usuario)

    def procesar_archivo(self, ruta, usuario):
        nombre_archivo = ruta.name
        
        # Ignorar archivos temporales
        if nombre_archivo.startswith('~$'):
            return
        
        match = re.match(r'^(\d{3})', nombre_archivo)
        if not match:
            self.stdout.write(self.style.WARNING(f'No puedo extraer código de {nombre_archivo}'))
            return

        codigo_centro = match.group(1)

        centro = Centros.objects.filter(codigo=codigo_centro).first()
        if centro is None:
            self.stderr.write(f'No existe Centro con código {codigo_centro}')
            return

        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        total_seguimentos = 0
        
        if 'Centro' in wb.sheetnames:
            cursos = detectar_cursos(wb['Centro'])
            if not cursos:
                self.stdout.write(self.style.WARNING(f'{nombre_archivo}: No se detectaron cursos'))
            else:
                n_seg = self.procesar_hoja(wb['Centro'], centro=centro, titulo=None, usuario=usuario, cursos=cursos)
                total_seguimentos += n_seg
        

        # --- Hoja "Centro" ---
        if 'Centro' in wb.sheetnames:
            cursos = detectar_cursos(wb['Centro'])
            n_seg = self.procesar_hoja(wb['Centro'], centro=centro, titulo=None, usuario=usuario, cursos=cursos)
            total_seguimentos += n_seg

        # --- Hojas de título ---
        for nombre_hoja in wb.sheetnames:
            if nombre_hoja in SHEETS_EXCLUIDAS or nombre_hoja == 'Centro':
                continue

            titulo = None
            for t in centro.titulos.all():
                ws = wb[nombre_hoja]
                for fila in ws.iter_rows(max_row=5, values_only=True):
                    for celda in fila:
                        if celda and isinstance(celda, str) and t.denominacion in celda:
                            titulo = t
                            break
                    if titulo:
                        break
                if titulo:
                    break

            if titulo is None:
                continue

            cursos = detectar_cursos(wb[nombre_hoja])
            n_seg = self.procesar_hoja(wb[nombre_hoja], centro=centro, titulo=titulo, usuario=usuario, cursos=cursos)
            total_seguimentos += n_seg

        self.stdout.write(self.style.SUCCESS(
            f'{nombre_archivo}: {total_seguimentos} seguimentos'
        ))

    def procesar_hoja(self, ws, centro, titulo, usuario, cursos):
        n_seguimentos = 0
        
        for fila in ws.iter_rows(min_row=7, values_only=True):
            if celda_segura(fila, 0) is None:
                continue

            codigo = str(celda_segura(fila, 0)).strip()
            
            # Saltar filas de categoría
            if 'Indicadores' in codigo:
                continue


            indicador = Indicadores.objects.filter(codigo=codigo).first()
            if indicador is None:
                self.stdout.write(self.style.WARNING(f'  Indicador {codigo} no encontrado'))
                continue

            

            for orixe_datos, (col_meta, col_resultado, col_obs) in cursos.items():
                meta = valor_seguimento(celda_segura(fila, col_meta - 1))
                resultado = valor_seguimento(celda_segura(fila, col_resultado - 1))
                observacions = limpiar_valor(celda_segura(fila, col_obs - 1))

                if meta is None and resultado is None and not observacions:
                    continue

                tipo_meta = 'porcentaje' if (meta is not None and 0 <= meta <= 1) else 'numero'

                if titulo:
                    seguimiento, creado = SeguimentosTitulos.objects.get_or_create(
                        titulo=titulo,
                        indicador=indicador,
                        orixe_datos=orixe_datos,
                        defaults={
                            'meta': meta,
                            'tipo_meta': tipo_meta,
                            'resultado': resultado,
                            'observacions': observacions,
                            'creado_por': usuario,
                        }
                    )
                else:
                    seguimiento, creado = Seguimentos.objects.get_or_create(
                        centro=centro,
                        indicador=indicador,
                        orixe_datos=orixe_datos,
                        defaults={
                            'meta': meta,
                            'tipo_meta': tipo_meta,
                            'resultado': resultado,
                            'observacions': observacions,
                            'creado_por': usuario,
                        }
                    )
                if creado:
                    n_seguimentos += 1

        return n_seguimentos
    