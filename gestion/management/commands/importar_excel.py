import re
from decimal import Decimal, InvalidOperation
from datetime import date

import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gestion.models import (
    IRPD, Centros, Titulos, Indicadores, Seguimentos, SeguimentosTitulos
)

User = get_user_model()

# Mapeo: nombre de la hoja -> denominación del título (tal y como está en tu tabla Titulos)
HOJAS_TITULOS = {
    'GEnxElectri': 'Grao en Enxeñaría Eléctrica',
    'GEnxElectro': 'Grao en Enxeñaría en Electrónica Industrial e Automática',
    'GEnxOrg': 'Grao en Enxeñaría en Organización Industrial',
    'GEnxQuim': 'Grao en Enxeñaría en Química Industrial',
    'GEnxTecInd': 'Grao en Enxeñaría en Tecnoloxías Industriais',
    'GEnxMec': 'Grao en Enxeñaría Mecánica',
    'GEnxBiomed': 'Grao en Enxeñaría Biomédica',
    'PCEOBioElectro': 'PCEO Grao en Enx. Biomédica/Grao en Enx. en Electrónica Industrial e Automática',
    'PCEOBioMec': 'PCEO Grao en Enx. Biomédica/Grao en Enx. Mecánica',
    'PCEOElectroMec': 'PCEO Grao en Enx. Mecánica/Grao en Enx. en Electrónica Industrial e Automática',
}

CURSOS_COLUMNAS = [
    # (columna_meta, columna_resultado, columna_observacions, curso_academico)
    (6, 7, 8, '23/24'),
    (10, 11, 12, '24/25'),
    (14, 15, 16, '25/26'),
]

REGEX_URL = re.compile(r'https?://\S+')

def celda_segura(fila, indice):
    """Devuelve fila[indice] o None si esa columna no existe en la fila."""
    if indice < len(fila):
        return fila[indice]
    return None

def limpiar_valor(valor):
    """Convierte '----', None, '' en None. Resto lo deja igual."""
    if valor is None:
        return None
    if isinstance(valor, str) and valor.strip() in ('----', '', '-'):
        return None
    return valor


def a_decimal(valor):
    """Intenta convertir a Decimal. Si no puede (texto raro), devuelve None."""
    valor = limpiar_valor(valor)
    if valor is None:
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        return None


def extraer_url(texto):
    if not texto:
        return None
    match = REGEX_URL.search(texto)
    return match.group(0) if match else None


def es_fila_categoria(fila):
    """Las filas de categoría solo tienen texto en la primera columna, el resto vacío."""
    return fila[0] is not None and all(c is None for c in fila[1:5])


class Command(BaseCommand):
    help = 'Importa indicadores y seguimentos desde el Excel del cadro de mando'

    def add_arguments(self, parser):
        parser.add_argument('ruta_excel', type=str)
        parser.add_argument('codigo_centro', type=str, help="Código del centro, ej: 312")

    def handle(self, *args, **options):
        ruta = options['ruta_excel']
        codigo_centro = options['codigo_centro']

        usuario_importacion = User.objects.filter(is_superuser=True).first()
        if usuario_importacion is None:
            self.stderr.write('No hay ningún superusuario para asignar como creado_por. Crea uno primero.')
            return

        centro = Centros.objects.filter(codigo=codigo_centro).first()
        if centro is None:
            self.stderr.write(f'No existe ningún Centro con código {codigo_centro}')
            return

        wb = openpyxl.load_workbook(ruta, data_only=True)

        total_indicadores = 0
        total_seguimentos = 0

        # --- Hoja "Centro" ---
        if 'Centro' in wb.sheetnames:
            n_ind, n_seg = self.procesar_hoja(
                wb['Centro'], centro=centro, titulo=None, usuario=usuario_importacion
            )
            total_indicadores += n_ind
            total_seguimentos += n_seg
            self.stdout.write(f'Hoja Centro: {n_ind} indicadores, {n_seg} seguimentos')

        # --- Hojas de título ---
        for nombre_hoja, denominacion_titulo in HOJAS_TITULOS.items():
            if nombre_hoja not in wb.sheetnames:
                continue
            titulo = Titulos.objects.filter(
                centro=centro, denominacion=denominacion_titulo
            ).first()
            if titulo is None:
                self.stdout.write(self.style.WARNING(
                    f'Aviso: no existe el Título "{denominacion_titulo}" en {centro}. Salto la hoja {nombre_hoja}.'
                ))
                continue

            n_ind, n_seg = self.procesar_hoja(
                wb[nombre_hoja], centro=centro, titulo=titulo, usuario=usuario_importacion
            )
            total_indicadores += n_ind
            total_seguimentos += n_seg
            self.stdout.write(f'Hoja {nombre_hoja}: {n_ind} indicadores, {n_seg} seguimentos')

        self.stdout.write(self.style.SUCCESS(
            f'Importación terminada. Indicadores tocados: {total_indicadores}. Seguimentos creados: {total_seguimentos}.'
        ))

    def procesar_hoja(self, ws, centro, titulo, usuario):
        n_indicadores = 0
        n_seguimentos = 0

        for fila in ws.iter_rows(min_row=6, values_only=True):
            if fila[0] is None:
                continue  # fila en blanco
            if es_fila_categoria(fila):
                continue  # fila de título de categoría, no es un indicador

            codigo = str(celda_segura(fila, 0)).strip()
            denominacion_indicador = (celda_segura(fila, 1) or '').strip()
            denominacion = (celda_segura(fila, 2) or '').strip()
            descricion = (celda_segura(fila, 3) or '').strip()
            irpd_celda = celda_segura(fila, 4)
            fonte = extraer_url(descricion)

            criterios = self.obtener_criterios(irpd_celda)

            for criterio in criterios:
                indicador = self.obtener_o_crear_indicador(
                    codigo, denominacion_indicador, denominacion,
                    descricion, fonte, criterio, usuario
                )
                n_indicadores += 1

                # Asegurar el vínculo N:M con el centro y, si procede, el título
                indicador.centros.add(centro)
                if titulo:
                    indicador.titulos.add(titulo)

                for col_meta, col_resultado, col_obs, curso in CURSOS_COLUMNAS:
                    meta = a_decimal(celda_segura(fila, col_meta))
                    resultado = a_decimal(celda_segura(fila, col_resultado))
                    observacions = limpiar_valor(celda_segura(fila, col_obs))

                    if meta is None and resultado is None:
                        continue  # nada que guardar para este curso

                    tipo_meta = 'porcentaje' if (meta is not None and 0 <= meta <= 1) else 'numero'

                    if titulo:
                        seguimiento, creado = SeguimentosTitulos.objects.get_or_create(
                            titulo=titulo,
                            indicador=indicador,
                            curso_academico=curso,
                            defaults={
                                'meta': meta,
                                'tipo_meta': tipo_meta,
                                'resultado': resultado,
                                'observacions': observacions,
                                'creado_por': usuario,
                            }
                        )
                    else:
                        seguimiento, creado = Seguimentos.objects.get_or_create(
                            centro=centro,
                            indicador=indicador,
                            curso_academico=curso,
                            defaults={
                                'meta': meta,
                                'tipo_meta': tipo_meta,
                                'resultado': resultado,
                                'observacions': observacions,
                                'creado_por': usuario,
                            }
                        )
                    if creado:
                        n_seguimentos += 1

        return n_indicadores, n_seguimentos

    def obtener_criterios(self, irpd_celda):
        """Convierte '1 e 7' en [1, 7]. Si está vacío, devuelve [None]."""
        if irpd_celda is None:
            return [None]
        texto = str(irpd_celda)
        numeros = re.findall(r'\d+', texto)
        if not numeros:
            return [None]
        return [int(n) for n in numeros]

    def obtener_o_crear_indicador(self, codigo, denominacion_indicador, denominacion,
                                descricion, fonte, criterio, usuario):
        irpd_obj = None
        if criterio is not None:
            irpd_obj, _ = IRPD.objects.get_or_create(
                criterio=str(criterio),
                defaults={'denominacion': f'Criterio {criterio}', 'creado_por': usuario}
            )
        else:
            # Si no hay criterio asignado, usar el IRPD por defecto "8 - Sin clasificar"
            irpd_obj, _ = IRPD.objects.get_or_create(
                criterio='8',
                defaults={'denominacion': 'Sin clasificar', 'creado_por': usuario}
            )

        # Si hay varios criterios para el mismo código, distinguimos el código para no chocar
        codigo_final = codigo
        if criterio is not None:
            existentes = Indicadores.objects.filter(codigo=codigo).exclude(irpd=irpd_obj)
            if existentes.exists():
                codigo_final = f"{codigo}-{criterio}"

        indicador, creado = Indicadores.objects.get_or_create(
            codigo=codigo_final,
            defaults={
                'denominacion_indicador': denominacion_indicador,
                'denominacion': denominacion,
                'descricion': descricion,
                'fonte': fonte,
                'irpd': irpd_obj,
                'sentido': 'positivo',  # valor por defecto, hay que revisarlo a mano después
                'creado_por': usuario,
            }
        )
        return indicador