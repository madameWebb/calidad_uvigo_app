import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from gestion.models import Codigos, Titulos

User = get_user_model()


class Command(BaseCommand):
    help = 'Importa códigos desde el Excel de titulaciones'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta del archivo Excel')

    def traducir_a_gallego(self, nombre):
        """Traduce términos castellanos comunes a gallego"""
        traducciones = {
            'Ambientales': 'Ambientais',
            'Ingeniería': 'Enxeñaría',
            'Enfermería': 'Enfermaría',
            'Tecnología': 'Tecnoloxía',
            'Arqueología': 'Arqueoloxía',
            'Ciudades': 'Cidades',
            'Inteligencia': 'Intelixencia',
            'Bachillerato': 'Bacharelato',
            'Biotecnología': 'Biotecnoloxía',
            'Biología': 'Bioloxía', 
            'Abordaje': 'Abordaxe',
            'Genética': 'Xenética',
            'PCEO': 'PCEO',
            'Energía': 'Enerxía',
            'Industriales': 'Industriais',
            'Extranjera': 'Extranxeira',
            'Realidad': 'Realidade',
            'Geoespacial': 'Xeoespacial',
            'Literatura': 'Literatura',
            'Música': 'Música',
            'Gestión': 'Xestión',
            'Desarrollo': 'Desenvolvemento',
            'Sostenible': 'Sostible',
            'Grado': 'Grao',
            'Máster': 'Máster',
            'Universitario': 'Universitario',
            'Nanotecnología': 'Nanotecnoloxía',
            'Institucional': 'Institucional',
            'Administración': 'Administración',
            'Dirección': 'Dirección',
            'Empresas': 'Empresas',
            'Tecnologías': 'Tecnoloxías',
        }
        
        nombre_traducido = nombre.lower()
        for castellano, gallego in traducciones.items():
            nombre_traducido = nombre_traducido.replace(castellano, gallego)
        
        return nombre_traducido

    def buscar_titulo(self, nombre):
        """Busca un título por palabras clave (2+ coincidencias)"""
        nombre_traducido = self.traducir_a_gallego(nombre)
        palabras_genéricas = {'en', 'de', 'y', 'e''Grado', 'Grao', 'Máster', 'Universitario'}
        palabras_clave = [p for p in nombre_traducido.lower().split() 
                          if len(p) > 3 and p not in palabras_genéricas]
        
        titulos = Titulos.objects.all()
        mejor_match = None
        mejor_coincidencias = 0
        
        for titulo in titulos:
            denominacion = titulo.denominacion.lower()
            coincidencias = sum(1 for palabra in palabras_clave if palabra in denominacion)
            
            if coincidencias > mejor_coincidencias:
                mejor_coincidencias = coincidencias
                mejor_match = titulo
        
        if mejor_coincidencias >= 1:
            return mejor_match
        
        return None

    def handle(self, *args, **options):
        archivo = Path(options['archivo'])

        if not archivo.exists():
            self.stderr.write(f'Archivo no encontrado: {archivo}')
            return

        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stderr.write('No hay superusuario.')
            return

        wb = openpyxl.load_workbook(str(archivo), data_only=True)
        ws = wb.active
        
        n_codigos = 0
        n_errores = 0
        n_no_encontrados = 0
        filas_no_importadas = []

        # Leer encabezado
        encabezado = [cell.value for cell in ws[1]]

        # Leer datos desde fila 2
        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                plan_sigma = fila[0]
                estudio_sigma = fila[1]
                xescampus = fila[5]
                nombre = fila[6]
                ruct = fila[7]

                if not plan_sigma or not nombre:
                    continue

                titulo = self.buscar_titulo(nombre)
                if not titulo:
                    self.stdout.write(self.style.WARNING(f'Fila {fila_num}: "{nombre}" no encontrado'))
                    filas_no_importadas.append(fila)
                    n_no_encontrados += 1
                    continue
                
                print(f"DEBUG Fila {fila_num}: Buscando '{nombre}'")
                titulo = self.buscar_titulo(nombre)
                print(f"DEBUG Resultado: {titulo}")

                if not titulo:
                    self.stdout.write(self.style.WARNING(f'Fila {fila_num}: "{nombre}" no encontrado'))
                    filas_no_importadas.append(fila)
                    n_no_encontrados += 1
                    continue

                codigo, creado = Codigos.objects.get_or_create(
                    plan_sigma=str(plan_sigma).strip() if plan_sigma else '',
                    defaults={
                        'titulo': titulo,
                        'estudio_sigma': str(estudio_sigma).strip() if estudio_sigma else '',
                        'xescampus': str(xescampus).strip() if xescampus else '',
                        'ruct': str(ruct).strip() if ruct else '',
                        'creado_por': usuario,
                    }
                )

                n_codigos += 1

            except Exception as e:
                self.stderr.write(f'Fila {fila_num}: Error - {e}')
                n_errores += 1

        # Crear Excel con filas no importadas
        if filas_no_importadas:
            wb_no_importadas = openpyxl.Workbook()
            ws_no_importadas = wb_no_importadas.active
            
            # Escribir encabezado
            for col_num, valor in enumerate(encabezado, 1):
                ws_no_importadas.cell(row=1, column=col_num, value=valor)
            
            # Escribir filas no importadas
            for row_num, fila in enumerate(filas_no_importadas, 2):
                for col_num, valor in enumerate(fila, 1):
                    ws_no_importadas.cell(row=row_num, column=col_num, value=valor)
            
            archivo_pendientes = Path(f'{archivo.stem}_pendientes.xlsx')
            wb_no_importadas.save(str(archivo_pendientes))
            self.stdout.write(self.style.SUCCESS(f'Archivo de pendientes: {archivo_pendientes}'))

        self.stdout.write(self.style.SUCCESS(
            f'{n_codigos} códigos importados, {n_no_encontrados} no encontrados, {n_errores} errores'
        ))