import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gestion.models import Centros, Titulos, Indicadores, Seguimentos, SeguimentosTitulos, Localizadores

User = get_user_model()

SHEETS_EXCLUIDAS = {'Portada', 'Centro', 'Anexos 1', 'Anexos 2', 'Anexos 3', 'Resumo', 'Procedementos'}
CURSOS_COLUMNAS = [
    (6, 7, 8, '23/24'),
    (10, 11, 12, '24/25'),
    (14, 15, 16, '25/26'),
]


def celda_segura(fila, indice):
    if indice < len(fila):
        return fila[indice]
    return None


def limpiar_valor(valor):
    if valor is None:
        return None
    if isinstance(valor, str) and valor.strip() in ('----', '', '-'):
        return None
    return valor


def a_decimal(valor):
    valor = limpiar_valor(valor)
    if valor is None:
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = 'Importa seguimentos desde los Excel (asume que los indicadores ya existen)'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        
        usuario = User.objects.filter(is_superuser=True).first()
        if usuario is None:
            self.stderr.write('No hay superusuario para asignar como creado_por.')
            return
        
        archivos = list(carpeta.glob('*.xlsx'))
        if not archivos:
            self.stderr.write(f'No hay archivos .xlsx en {carpeta}')
            return

        for ruta in archivos:
            self.procesar_archivo(ruta, usuario)

    def procesar_archivo(self, ruta, usuario):
        nombre_archivo = ruta.name
        match = re.match(r'^(\d)(\d+)', nombre_archivo)
        if not match:
            self.stdout.write(self.style.WARNING(f'No puedo extraer código de {nombre_archivo}'))
            return
        
        codigo_localizador = match.group(1)
        codigo_centro = match.group(2)

        localizador = Localizadores.objects.filter(codigo=codigo_localizador).first()

        if localizador is None:
            self.stderr.write(f'Localizador {codigo_localizador} no existe.')
            return

        centro = Centros.objects.filter(codigo=codigo_centro, codigo_localizador=localizador).first()
        if centro is None:
            self.stderr.write(f'No existe Centro con código {codigo_centro}')
            return

        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        total_seguimentos = 0

        # --- Hoja "Centro" ---
        if 'Centro' in wb.sheetnames:
            n_seg = self.procesar_hoja(wb['Centro'], centro=centro, titulo=None, usuario=usuario)
            total_seguimentos += n_seg

        # --- Hojas de título ---
        for nombre_hoja in wb.sheetnames:
            if nombre_hoja in SHEETS_EXCLUIDAS:
                continue

            titulo = None
            for t in centro.titulos.all():
                ws = wb[nombre_hoja]
                for fila in ws.iter_rows(max_row=5, values_only=True):
                    for celda in fila:
                        if celda and isinstance(celda, str) and t.denominacion in celda:
                            titulo = t
                            break
                    if titulo:
                        break
                if titulo:
                    break

            if titulo is None:
                continue

            n_seg = self.procesar_hoja(wb[nombre_hoja], centro=centro, titulo=titulo, usuario=usuario)
            total_seguimentos += n_seg

        self.stdout.write(self.style.SUCCESS(
            f'{nombre_archivo}: {total_seguimentos} seguimentos'
        ))

    def procesar_hoja(self, ws, centro, titulo, usuario):
        n_seguimentos = 0

        for fila in ws.iter_rows(min_row=6, values_only=True):
            if celda_segura(fila, 0) is None:
                continue

            codigo = str(celda_segura(fila, 0)).strip()

            indicador = Indicadores.objects.filter(codigo=codigo).first()
            if indicador is None:
                continue

            for col_meta, col_resultado, col_obs, curso in CURSOS_COLUMNAS:
                meta = a_decimal(celda_segura(fila, col_meta))
                resultado = a_decimal(celda_segura(fila, col_resultado))
                observacions = limpiar_valor(celda_segura(fila, col_obs))

                if meta is None and resultado is None:
                    continue

                tipo_meta = 'porcentaje' if (meta is not None and 0 <= meta <= 1) else 'numero'

                if titulo:
                    seguimiento, creado = SeguimentosTitulos.objects.get_or_create(
                        titulo=titulo,
                        indicador=indicador,
                        orixe_datos=curso,
                        defaults={
                            'meta': meta,
                            'tipo_meta': tipo_meta,
                            'resultado': resultado,
                            'observacions': observacions,
                            'creado_por': usuario,
                        }
                    )
                else:
                    seguimiento, creado = Seguimentos.objects.get_or_create(
                        centro=centro,
                        indicador=indicador,
                        orixe_datos=curso,
                        defaults={
                            'meta': meta,
                            'tipo_meta': tipo_meta,
                            'resultado': resultado,
                            'observacions': observacions,
                            'creado_por': usuario,
                        }
                    )
                if creado:
                    n_seguimentos += 1

        return n_seguimentos