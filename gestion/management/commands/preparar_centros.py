import re 
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gestion.models import Localizadores, Centros, Titulos

User = get_user_model()


def extraer_codigo_de_nombre(nombre_archivo):
    """Del nombre '312_DE-02_Cadro...' extrae localizador='3', centro='12'."""
    match = re.match(r'^(\d+)', nombre_archivo)
    if not match:
        return None, None
    codigo_completo = match.group(1)
    return codigo_completo[0], codigo_completo[1:]


def extraer_nombre_centro(ws_portada, nombre_archivo=None):
    """Intento 1: buscar en Portada. Intento 2: deducir del nombre de archivo."""
    if ws_portada is not None:
        for fila in ws_portada.iter_rows(values_only=True):
            for celda in fila:
                if celda and isinstance(celda, str) and 'Cadro de Mando' in celda:
                    texto = re.sub(
                        r'^.*Cadro de Mandos?\s+(d[oa]s?)\s+(Graos?|Másteres?)\s+(d[oa]s?)\s+',
                        '', celda
                    )
                    texto = texto.strip()
                    if texto and texto != celda.strip():
                        return texto

    if nombre_archivo:
        # quita extensión, código inicial, y palabras técnicas conocidas
        texto = Path(nombre_archivo).stem
        texto = re.sub(r'^\d+_?', '', texto)            # quita el código inicial
        texto = re.sub(r'DE-?\d*', '', texto)             # quita 'DE-02'
        texto = re.sub(r'Cadro_de_Mando[s]?', '', texto)  # quita 'Cadro_de_Mando'
        texto = re.sub(r'\d{2}-\d{2}', '', texto)         # quita '25-26'
        texto = re.sub(r'Graos?|Másteres?|Modificado', '', texto)
        texto = texto.replace('_', ' ').strip()
        return texto if texto else None

    return None


def extraer_titulos(wb, sheetnames_excluidas):
    titulos = []
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja in sheetnames_excluidas:
            continue
        ws = wb[nombre_hoja]
        for fila in ws.iter_rows(max_row=5, values_only=True):
            for celda in fila:
                if celda and isinstance(celda, str) and 'Cadro de mando' in celda:
                    match = re.search(r'Cadro de mando\s+d[oa]s?\s+(.+)', celda)
                    if match:
                        denominacion = match.group(1).strip()
                        tipo = 'master' if denominacion.startswith('M.') or 'Máster' in denominacion or 'M. Univ' in denominacion else 'grado'
                        titulos.append((nombre_hoja, denominacion, tipo))
                        break
            else:
                continue
            break
    return titulos


SHEETS_EXCLUIDAS = {'Portada', 'Centro', 'Anexos 1', 'Anexos 2', 'Anexos 3', 'Resumo', 'Procedementos'}


class Command(BaseCommand):
    help = 'Crea Centros y Titulos a partir de una carpeta con los Excel de cadro de mando'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str)

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        usuario = User.objects.filter(is_superuser=True).first()
        if usuario is None:
            self.stderr.write('No hay superusuario disponible para creado_por.')
            return

        for archivo in carpeta.glob('*.xlsx'):
            cod_localizador, cod_centro = extraer_codigo_de_nombre(archivo.name)
            if cod_centro is None:
                self.stdout.write(self.style.WARNING(f'No pude leer el código en {archivo.name}, lo salto.'))
                continue

            localizador = Localizadores.objects.filter(codigo=cod_localizador).first()
            if localizador is None:
                self.stdout.write(self.style.WARNING(
                    f'No existe Localizador con código {cod_localizador} (archivo {archivo.name}). Lo salto.'
                ))
                continue

            wb = openpyxl.load_workbook(archivo, data_only=True)

            nombre_centro = None
            if 'Portada' in wb.sheetnames:
                nombre_centro = extraer_nombre_centro(wb['Portada'], archivo.name)
            else:
                nombre_centro = extraer_nombre_centro(None, archivo.name)

            if not nombre_centro:
                self.stdout.write(self.style.WARNING(
                    f'No pude extraer el nombre del centro en {archivo.name}. Revísalo a mano.'
                ))
                continue

            centro, creado = Centros.objects.get_or_create(
                codigo=cod_centro,
                codigo_localizador=localizador,   # ← añadir esto
                defaults={
                    'denominacion': nombre_centro,
                    'creado_por': usuario,
                }
            )
            estado = 'creado' if creado else 'ya existía'
            self.stdout.write(f'Centro {centro.codigo} ({centro.denominacion}): {estado}')

            titulos_detectados = extraer_titulos(wb, SHEETS_EXCLUIDAS)
            for nombre_hoja, denominacion, tipo in titulos_detectados:
                titulo, creado_t = Titulos.objects.get_or_create(
                    centro=centro,
                    denominacion=denominacion,
                    defaults={'tipo': tipo, 'creado_por': usuario}
                )
                estado_t = 'creado' if creado_t else 'ya existía'
                self.stdout.write(f'  Título "{denominacion}" (hoja {nombre_hoja}): {estado_t}')