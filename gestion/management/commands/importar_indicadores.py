import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

from gestion.models import IRPD, Centros, Titulos, Indicadores, Localizadores

User = get_user_model()

REGEX_URL = re.compile(r'https?://\S+')
SHEETS_EXCLUIDAS = {'Portada', 'Anexos 1', 'Anexos 2', 'Anexos 3', 'Resumo', 'Procedementos'}


def extraer_url(texto):
    """Extrae URL de un texto."""
    if not texto:
        return None
    match = REGEX_URL.search(str(texto))
    return match.group(0) if match else None


def limpiar_valor(valor):
    """Convierte '----', None, '' en None."""
    if valor is None:
        return None
    if isinstance(valor, str) and valor.strip() in ('----', '', '-'):
        return None
    return valor


def obtener_tipo_indicador(celda):
    """Detecta el tipo de indicador por color de fila."""
    if not celda or not celda.fill:
        return 'calidade'  # default
    
    fill = celda.fill
    if not fill.start_color:
        return 'calidade'
    
    rgb = str(fill.start_color.rgb).upper()
    
    # Azul = calidade (default)
    # Verde = institucional
    # Rojo = estratexico
    if 'FF00B050' in rgb or '00B050' in rgb:  # Verde
        return 'institucional'
    elif 'FFFF0000' in rgb or 'FF0000' in rgb:  # Rojo
        return 'estratexico'
    
    return 'calidade'


class Command(BaseCommand):
    help = 'Importa indicadores desde Excel del cadro de mando'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stderr.write('No hay superusuario.')
            return

        archivos = list(carpeta.glob('*.xlsx'))
        if not archivos:
            self.stderr.write(f'No hay .xlsx en {carpeta}')
            return

        for ruta in archivos:
            self.procesar_archivo(ruta, usuario)

    def procesar_archivo(self, ruta, usuario):
        nombre_archivo = ruta.name
        match = re.match(r'^(\d)(\d+)', nombre_archivo)
        if not match:
            self.stdout.write(self.style.WARNING(f'No puedo extraer código de {nombre_archivo}'))
            return

        codigo_localizador = match.group(1)
        codigo_centro = match.group(2)

        localizador = Localizadores.objects.filter(codigo=codigo_localizador).first()
        if not localizador:
            self.stderr.write(f'Localizador {codigo_localizador} no existe.')
            return

        centro = Centros.objects.filter(codigo=codigo_centro, codigo_localizador=localizador).first()
        if not centro:
            self.stderr.write(f'Centro {codigo_centro} no existe.')
            return

        wb = openpyxl.load_workbook(str(ruta), data_only=False)  # data_only=False para detectar colores
        total_indicadores = 0

        # Hoja "Centro"
        if 'Centro' in wb.sheetnames:
            n = self.procesar_hoja(wb['Centro'], centro, None, usuario)
            total_indicadores += n

        # Hojas de títulos
        for nombre_hoja in wb.sheetnames:
            if nombre_hoja in SHEETS_EXCLUIDAS:
                continue

            titulo = self.buscar_titulo(wb[nombre_hoja], centro)
            if not titulo:
                continue

            n = self.procesar_hoja(wb[nombre_hoja], centro, titulo, usuario)
            total_indicadores += n

        self.stdout.write(self.style.SUCCESS(
            f'{nombre_archivo}: {total_indicadores} indicadores'
        ))

    def buscar_titulo(self, ws, centro):
        """Busca el título que corresponde a esta hoja."""
        for fila in ws.iter_rows(max_row=5, values_only=True):
            for celda in fila:
                if celda and isinstance(celda, str):
                    for t in centro.titulos.all():
                        if t.denominacion in celda:
                            return t
        return None

    def procesar_hoja(self, ws, centro, titulo, usuario):
        """Lee la hoja y crea indicadores."""
        n_indicadores = 0
        tipo_actual = 'calidade'

        for fila_num, fila in enumerate(ws.iter_rows(min_row=6), start=6):
            valor_primera = str(fila[0].value or '').strip()
            
            # Detectar cambio de tipo por texto
            if 'CALIDADE' in valor_primera.upper():
                tipo_actual = 'calidade'
                continue
            elif 'INSTITUCION' in valor_primera.upper():
                tipo_actual = 'institucional'
                continue
            elif 'ESTRATÉXICO' in valor_primera.upper() or 'ESTRATEXICO' in valor_primera.upper():
                tipo_actual = 'estratexico'
                continue
            
            # Saltar filas vacías
            if not valor_primera:
                continue

            # Es un indicador, procesar normalmente
            codigo = valor_primera if valor_primera else None

            # Salta fórmulas
            if codigo.startswith('='):
                continue


            denominacion = str(fila[1].value).strip() if fila[1].value else None
            procedemento = str(fila[2].value).strip() if fila[2].value else None
            descricion = str(fila[3].value).strip() if fila[3].value else None
            irpd_celda = fila[4].value if len(fila) > 4 else None

            if not codigo or not denominacion:
                continue

            fonte = extraer_url(descricion)
            irpd_obj = self.obtener_irpd(irpd_celda, usuario)

            indicador, creado = Indicadores.objects.get_or_create(
                codigo=codigo,
                defaults={
                    'denominacion': denominacion,
                    'procedemento_asociado': procedemento or '',
                    'descricion': descricion or '',
                    'fonte': fonte,
                    'irpd': irpd_obj,
                    'tipo_indicador': tipo_actual,  # ← USA EL TIPO DETECTADO
                    'sentido': 'positivo',
                    'creado_por': usuario,
                }
            )

            indicador.centros.add(centro)
            if titulo:
                indicador.titulos.add(titulo)

            n_indicadores += 1

        return n_indicadores

    def obtener_irpd(self, irpd_celda, usuario):
        """Obtiene o crea IRPD."""
        if not irpd_celda:
            irpd, _ = IRPD.objects.get_or_create(
                criterio='8',
                defaults={'denominacion': 'Sin clasificar', 'creado_por': usuario}
            )
            return irpd

        texto = str(irpd_celda)
        numeros = re.findall(r'\d+', texto)
        
        if not numeros:
            irpd, _ = IRPD.objects.get_or_create(
                criterio='8',
                defaults={'denominacion': 'Sin clasificar', 'creado_por': usuario}
            )
            return irpd

        criterio = numeros[0]
        irpd, _ = IRPD.objects.get_or_create(
            criterio=criterio,
            defaults={'denominacion': f'Criterio {criterio}', 'creado_por': usuario}
        )
        return irpd