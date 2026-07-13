import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from gestion.models import MateriasAvaliadas, Indicadores, SeguementoMaterias

User = get_user_model()

ANEXOS_ESTANDAR = {'Anexos 1', 'Anexos 2', 'Anexos 3'}

def detectar_anexos(wb):
    cursos = {}
    cursos_orden = ['23/24', '24/25', '25/26']
    anexos_encontrados = sorted([h for h in wb.sheetnames if h.strip() in ANEXOS_ESTANDAR])
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques

class Command(BaseCommand):
    help = 'Importa seguimentos de materias desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stderr.write('No hay superusuario.')
            return

        archivos = sorted([f for f in carpeta.glob('*.xlsx') if not f.name.startswith('~$')])
        if not archivos:
            self.stderr.write(f'No hay archivos .xlsx en {carpeta}')
            return

        errores = []
        for ruta in archivos:
            errores_archivo = self.procesar_archivo(ruta, usuario)
            errores.extend(errores_archivo)

        if errores:
            salida = Path('importar_seguimentos_materias_errores.txt')
            with open(salida, 'w', encoding='utf-8') as f:
                f.write("ERROS NA IMPORTACIÓN\n")
                f.write("=" * 100 + "\n\n")
                for error in errores:
                    f.write(f"{error}\n")
            self.stdout.write(self.style.WARNING(f'Erros en: {salida}'))

    def procesar_archivo(self, ruta, usuario):
        errores = []
        nombre = ruta.name
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        cursos = detectar_anexos(wb)
        total = 0

        for nombre_hoja, orixe_datos in cursos.items():
            ws = wb[nombre_hoja]
            bloques = detectar_bloques(ws)

            # Leer códigos de indicadores (fila 2, primera col del bloque)
            for col_inicio in bloques:
                nombre_titulo = ws.cell(row=6, column=col_inicio).value

                # Saltar PCEO
                if nombre_titulo and str(nombre_titulo).upper().startswith('PCEO'):
                    continue

                # Leer indicadores de fila 2
                codigos_indicadores_raw = ws.cell(row=2, column=col_inicio).value
                if not codigos_indicadores_raw:
                    continue

                codigos_indicadores = [c.strip() for c in str(codigos_indicadores_raw).split('\n') if c.strip()]

                indicadores = []
                for codigo in codigos_indicadores:
                    ind = Indicadores.objects.filter(codigo=codigo).first()
                    if ind:
                        indicadores.append(ind)
                    else:
                        errores.append(f'{nombre} - {nombre_hoja}: Indicador {codigo} non encontrado')

                if not indicadores:
                    continue

                # Procesar filas de datos
                for fila_num in range(6, ws.max_row + 1):
                    codigo_materia = ws.cell(row=fila_num, column=col_inicio + 1).value

                    if not codigo_materia:
                        break

                    materia = MateriasAvaliadas.objects.filter(codigo=str(codigo_materia).strip()).first()
                    if not materia:
                        errores.append(f'{nombre} - {nombre_hoja} fila {fila_num}: Materia {codigo_materia} non encontrada')
                        continue

                    # Taxa indicador 1 → col+4, indicador 2 → col+5
                    # Meta indicador 1 → col+6, indicador 2 → col+7
                    taxas = [
                        ws.cell(row=fila_num, column=col_inicio + 4).value,
                        ws.cell(row=fila_num, column=col_inicio + 5).value,
                    ]
                    metas = [
                        ws.cell(row=fila_num, column=col_inicio + 6).value,
                        ws.cell(row=fila_num, column=col_inicio + 7).value,
                    ]

                    for i, indicador in enumerate(indicadores):
                        taxa = taxas[i] if i < len(taxas) else None
                        meta = metas[i] if i < len(metas) else None

                        # Convertir 'Sen Meta' a None
                        if isinstance(meta, str):
                            meta = None
                        if isinstance(taxa, str):
                            taxa = None

                        seg, creado = SeguementoMaterias.objects.get_or_create(
                            materia=materia,
                            indicador=indicador,
                            orixe_datos=orixe_datos,
                            defaults={
                                'taxa': taxa,
                                'meta': meta,
                                'creado_por': usuario,
                            }
                        )

                        if creado:
                            total += 1

        self.stdout.write(self.style.SUCCESS(f'{nombre}: {total} seguimentos importados'))
        return errores