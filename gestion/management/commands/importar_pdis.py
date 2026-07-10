import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from gestion.models import Indicadores, Titulos, Codigos, AvaliacionsPdis

User = get_user_model()

ANEXOS_ESTANDAR = {'Anexos 1', 'Anexos 2', 'Anexos 3'}

def detectar_anexos(wb):
    cursos = {}
    cursos_orden = ['23/24', '24/25', '25/26']
    anexos_encontrados = sorted([h for h in wb.sheetnames 
                                  if h.strip() in ANEXOS_ESTANDAR])
    for i, nombre_hoja in enumerate(anexos_encontrados):
        if i < len(cursos_orden):
            cursos[nombre_hoja] = cursos_orden[i]
    return cursos

class Command(BaseCommand):
    help = 'Importa avaliacións de PDIs desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stderr.write('No hay superusuario.')
            return

        archivos = [f for f in carpeta.glob('*.xlsx') if not f.name.startswith('~$')]
        if not archivos:
            self.stderr.write(f'No hay archivos .xlsx en {carpeta}')
            return

        errores = []
        
        for ruta in archivos:
            errores_archivo = self.procesar_archivo(ruta, usuario)
            errores.extend(errores_archivo)

        # Guardar errores en archivo
        if errores:
            salida = Path('importar_pdis_errores.txt')
            with open(salida, 'w', encoding='utf-8') as f:
                f.write("ERROS NA IMPORTACIÓN DE PDIs\n")
                f.write("=" * 100 + "\n\n")
                for error in errores:
                    f.write(f"{error}\n")
            self.stdout.write(self.style.WARNING(f'Erros gardados en: {salida}'))
            
    def procesar_archivo(self, ruta, usuario):
        errores = []
        nombre = ruta.name
        
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        cursos = detectar_anexos(wb)  # ← En lugar de ANEXOS_CURSOS
        
        if not cursos:
            errores.append(f'{nombre}: Non se atoparon Anexos estándar')
            return errores
        
        total = 0
       
        for nombre_hoja, orixe_datos in cursos.items():
            if nombre_hoja not in wb.sheetnames:
                errores.append(f'{nombre} - Hoja {nombre_hoja} ({orixe_datos}) non encontrada')
                continue
            
            ws = wb[nombre_hoja]

            # Fila 2: código del indicador
            codigo_indicador = ws.cell(row=2, column=2).value
            
            if not codigo_indicador:
                continue

            indicador = Indicadores.objects.filter(codigo=str(codigo_indicador).strip()).first()
            
            if not indicador:
                self.stdout.write(self.style.WARNING(f'{nombre}: Indicador {codigo_indicador} no encontrado'))
                errores.append(f'{ruta.name} - {nombre_hoja}: Indicador {codigo_indicador} non encontrado')
                continue
            
            # Fila 6+: datos por titulación
            for fila_num in range(6, ws.max_row + 1):
                xescampus = ws.cell(row=fila_num, column=2).value
                nombre_titulo = ws.cell(row=fila_num, column=3).value
                
                if not xescampus and not nombre_titulo:
                    break

                codigo_obj = Codigos.objects.filter(xescampus=str(xescampus).strip()).first() if xescampus else None

                if not codigo_obj:
                    if nombre_titulo:
                        titulo_obj = Titulos.objects.filter(denominacion__icontains=nombre_titulo).first()
                        if not titulo_obj:
                            errores.append(f'{nombre} - {nombre_hoja} fila {fila_num}: {xescampus} / "{nombre_titulo}" non encontrado')
                            continue
                    else:
                        errores.append(f'{nombre} - {nombre_hoja} fila {fila_num}: {xescampus} sen nome de título')
                        continue
                else:
                    titulo_obj = codigo_obj.titulo

                excelentes = ws.cell(row=fila_num, column=4).value or 0
                notables = ws.cell(row=fila_num, column=5).value or 0
                favorables = ws.cell(row=fila_num, column=6).value or 0
                desfavorables = ws.cell(row=fila_num, column=7).value or 0
                totales = ws.cell(row=fila_num, column=8).value or 0

                # Buscar título por xescampus
                codigo_obj = Codigos.objects.filter(xescampus=str(xescampus).strip()).first()
                if not codigo_obj:
                    self.stdout.write(self.style.WARNING(f'{nombre} fila {fila_num}: xescampus {xescampus} no encontrado'))
                    continue

                titulo = codigo_obj.titulo

                avaliacion, creado = AvaliacionsPdis.objects.get_or_create(
                    titulo=titulo,
                    orixe_datos=orixe_datos,
                    defaults={
                        'excelentes': excelentes,
                        'notables': notables,
                        'favorables': favorables,
                        'desfavorables': desfavorables,
                        'totales': totales,
                        'creado_por': usuario,
                    }
                )

                # Asignar indicador después
                if creado:
                    avaliacion.indicador.set([indicador])
                    total += 1

        self.stdout.write(self.style.SUCCESS(f'{nombre}: {total} avaliacións importadas'))
        return errores