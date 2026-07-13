import re
import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from gestion.models import Titulos, Centros, MateriasAvaliadas

User = get_user_model()

def traducir_a_gallego(nombre):
    traducciones = {
        'Ambientales': 'Ambientais',
        'Ingeniería': 'Enxeñaría',
        'Enfermería': 'Enfermaría',
        'Tecnología': 'Tecnoloxía',
        'Arqueología': 'Arqueoloxía',
        'Ciudades': 'Cidades',
        'Inteligencia': 'Intelixencia',
        'Bachillerato': 'Bacharelato',
        'Biotecnología': 'Biotecnoloxía',
        'Biología': 'Bioloxía',
        'Abordaje': 'Abordaxe',
        'Genética': 'Xenética',
        'Energía': 'Enerxía',
        'Industriales': 'Industriáis',
        'Extranjera': 'Extranxeira',
        'Extranjeras': 'Estranxeiras',
        'Realidad': 'Realidade',
        'Geoespacial': 'Xeoespacial',
        'Gestión': 'Xestión',
        'Desarrollo': 'Desenvolvemento',
        'Sostenible': 'Sostible',
        'Nanotecnología': 'Nanotecnoloxía',
        'Derecho': 'Dereito',
        'Diseño': 'Deseño',
        'Lenguas': 'Linguas',
        'Traducción': 'Tradución',
        'Filología': 'Filoloxía',
        'Lingüística': 'Lingüística',
        'Enseñanza': 'Ensino',
        'Relaciones': 'Relacións',
        'Internacionales': 'Internacionais'
    }
    nombre_traducido = nombre.lower()
    for castellano, gallego in traducciones.items():
        nombre_traducido = nombre_traducido.replace(castellano.lower(), gallego.lower())
    return nombre_traducido

def buscar_titulo(nombre, codigo_centro):
    if not nombre:
        return None
    
    nombre_traducido = traducir_a_gallego(nombre)
    palabras_genericas = {'en', 'de', 'y', 'e', 'grado', 'grao', 'máster', 'universitario', 'para', 'por', 'a'}
    palabras_clave = [p for p in nombre_traducido.lower().split()
                      if len(p) > 3 and p not in palabras_genericas]

    centro = Centros.objects.filter(codigo=codigo_centro).first()
    titulos = Titulos.objects.filter(centro=centro) if centro else Titulos.objects.all()

    mejor_match = None
    mejor_coincidencias = 0

    for titulo in titulos:
        denominacion = titulo.denominacion.lower()
        coincidencias = sum(1 for palabra in palabras_clave if palabra in denominacion)
        if coincidencias > mejor_coincidencias:
            mejor_coincidencias = coincidencias
            mejor_match = titulo

    return mejor_match if mejor_coincidencias >= 1 else None

def detectar_bloques(ws):
    bloques = []
    for col in range(10, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and 'Titulaci' in str(val):
            bloques.append(col)
    return bloques


class Command(BaseCommand):
    help = 'Importa materias avaliadas desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('carpeta', type=str, help='Carpeta con los Excel')

    def handle(self, *args, **options):
        carpeta = Path(options['carpeta'])
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stderr.write('No hay superusuario.')
            return

        archivos = sorted([f for f in carpeta.glob('*.xlsx') if not f.name.startswith('~$')])
        if not archivos:
            self.stderr.write(f'No hay archivos .xlsx en {carpeta}')
            return

        errores = []
        for ruta in archivos:
            errores_archivo = self.procesar_archivo(ruta, usuario)
            errores.extend(errores_archivo)

        if errores:
            salida = Path('importar_materias_errores.txt')
            with open(salida, 'w', encoding='utf-8') as f:
                f.write("ERROS NA IMPORTACIÓN DE MATERIAS\n")
                f.write("=" * 100 + "\n\n")
                for error in errores:
                    f.write(f"{error}\n")
            self.stdout.write(self.style.WARNING(f'Erros en: {salida}'))

    def procesar_archivo(self, ruta, usuario):
        errores = []
        nombre = ruta.name

        match = re.match(r'^(\d{3})', nombre)
        codigo_centro = match.group(1) if match else None

        wb = openpyxl.load_workbook(str(ruta), data_only=True)

        if 'Anexos 1' not in wb.sheetnames:
            errores.append(f'{nombre}: Non se atopou hoja Anexos 1')
            return errores

        ws = wb['Anexos 1']
        bloques = detectar_bloques(ws)
        total = 0

        for col_inicio in bloques:
            nombre_titulo = ws.cell(row=6, column=col_inicio).value
            titulo = buscar_titulo(nombre_titulo, codigo_centro) if nombre_titulo else None

            if not titulo:
                errores.append(f'{nombre} col {col_inicio}: Título "{nombre_titulo}" non encontrado')
                continue

            for fila_num in range(6, ws.max_row + 1):
                codigo_asig = ws.cell(row=fila_num, column=col_inicio + 1).value
                nome_asig = ws.cell(row=fila_num, column=col_inicio + 2).value

                if not codigo_asig and not nome_asig:
                    break

                if not codigo_asig or not nome_asig:
                    continue

                materia, creado = MateriasAvaliadas.objects.get_or_create(
                    codigo=str(codigo_asig).strip(),
                    defaults={
                        'materia': str(nome_asig).strip(),
                        'titulo': titulo,
                        'creado_por': usuario,
                    }
                )

                if creado:
                    total += 1

        self.stdout.write(self.style.SUCCESS(f'{nombre}: {total} materias importadas'))
        return errores