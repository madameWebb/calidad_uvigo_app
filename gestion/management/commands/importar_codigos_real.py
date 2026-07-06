import openpyxl
from pathlib import Path
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from gestion.models import Titulos, Codigos, Centros

User = get_user_model()


class Command(BaseCommand):
    help = 'Importa códigos desde el Excel de titulaciones'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta del archivo Excel')

    def traducir_a_gallego(self, nombre):
        traducciones = {
            'Ambientales': 'Ambientais',
            'Ingeniería': 'Enxeñaría',
            'Enfermería': 'Enfermaría',
            'Tecnología': 'Tecnoloxía',
            'Arqueología': 'Arqueoloxía',
            'Ciudades': 'Cidades',
            'Inteligencia': 'Intelixencia',
            'Bachillerato': 'Bacharelato',
            'Biotecnología': 'Biotecnoloxía',
            'Biología': 'Bioloxía',
            'Abordaje': 'Abordaxe',
            'Genética': 'Xenética',
            'Energía': 'Enerxía',
            'Industriales': 'Industriais',
            'Extranjera': 'Extranxeira',
            'Realidad': 'Realidade',
            'Geoespacial': 'Xeoespacial',
            'Gestión': 'Xestión',
            'Desarrollo': 'Desenvolvemento',
            'Sostenible': 'Sostible',
            'Nanotecnología': 'Nanotecnoloxía',
            'Administración': 'Administración',
            'Dirección': 'Dirección',
            'Empresas': 'Empresas',
            'Tecnologías': 'Tecnoloxías',
            'Derecho': 'Dereito',
            'Diseño': 'Deseño',
            'Extranjeras': 'Extranxeiras',
        }

        nombre_traducido = nombre.lower()
        for castellano, gallego in traducciones.items():
            nombre_traducido = nombre_traducido.replace(castellano.lower(), gallego.lower())

        return nombre_traducido

    def buscar_titulo(self, nombre, localizador):
        nombre_traducido = self.traducir_a_gallego(nombre)
        palabras_genericas = {'en', 'de', 'y', 'e', 'grado', 'grao', 'máster', 'universitario'}
        palabras_clave = [p for p in nombre_traducido.lower().split()
                          if len(p) > 3 and p not in palabras_genericas]

        centro = Centros.objects.filter(codigo__icontains=localizador).first()

        if centro:
            titulos = Titulos.objects.filter(centro=centro)
        else:
            titulos = Titulos.objects.all()

        mejor_match = None
        mejor_coincidencias = 0

        for titulo in titulos:
            denominacion = titulo.denominacion.lower()
            coincidencias = sum(1 for palabra in palabras_clave if palabra in denominacion)

            if coincidencias > mejor_coincidencias:
                mejor_coincidencias = coincidencias
                mejor_match = titulo

        if mejor_coincidencias >= 1:
            return mejor_match

        return None

    def handle(self, *args, **options):
        archivo = Path(options['archivo'])

        if not archivo.exists():
            self.stderr.write(f'Archivo no encontrado: {archivo}')
            return

        # Usuario del sistema para auditoría
        usuario = User.objects.filter(is_superuser=True).first()

        # Títulos que ya tienen código asignado
        titulos_con_codigo = set(Codigos.objects.values_list('titulo_id', flat=True))

        wb = openpyxl.load_workbook(str(archivo), data_only=True)
        ws = wb.active

        # Guardar encabezado para el Excel de pendientes
        encabezado = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        n_importados = 0
        n_saltados = 0
        n_no_encontrados = 0
        filas_no_importadas = []

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                plan_sigma    = fila[0]
                estudio_sigma = fila[1]
                localizador   = fila[3]
                xescampus     = fila[5]
                nombre        = fila[6]
                ruct          = fila[7]

                if not plan_sigma or not nombre:
                    continue

                titulo = self.buscar_titulo(nombre, localizador)

                if not titulo:
                    self.stdout.write(f'? Fila {fila_num}: "{nombre}" NO ENCONTRADO (localizador: {localizador})')
                    n_no_encontrados += 1
                    filas_no_importadas.append(fila)
                    continue

                if titulo.id in titulos_con_codigo:
                    self.stdout.write(f'✗ Fila {fila_num}: "{titulo.denominacion}" YA TIENE CÓDIGO')
                    n_saltados += 1
                    continue

                codigo, creado = Codigos.objects.get_or_create(
                    plan_sigma=str(plan_sigma).strip(),
                    defaults={
                        'titulo':         titulo,
                        'estudio_sigma':  str(estudio_sigma).strip() if estudio_sigma else '',
                        'xescampus':      str(xescampus).strip()     if xescampus     else '',
                        'ruct':           str(ruct).strip()           if ruct           else '',
                        'creado_por':     usuario,
                    }
                )

                if creado:
                    self.stdout.write(f'✓ Fila {fila_num}: "{titulo.denominacion}" importado')
                    n_importados += 1
                else:
                    self.stdout.write(f'✗ Fila {fila_num}: plan_sigma {plan_sigma} ya existía en Codigos')
                    n_saltados += 1

            except Exception as e:
                self.stderr.write(f'✗ Fila {fila_num}: Error - {e}')
                filas_no_importadas.append(fila)

        # Generar Excel con las filas que no se pudieron importar
        if filas_no_importadas:
            wb_pendientes = openpyxl.Workbook()
            ws_pendientes = wb_pendientes.active

            for col_num, valor in enumerate(encabezado, 1):
                ws_pendientes.cell(row=1, column=col_num, value=valor)

            for row_num, fila in enumerate(filas_no_importadas, 2):
                for col_num, valor in enumerate(fila, 1):
                    ws_pendientes.cell(row=row_num, column=col_num, value=valor)

            archivo_pendientes = archivo.parent / f'{archivo.stem}_pendientes.xlsx'
            wb_pendientes.save(str(archivo_pendientes))
            self.stdout.write(self.style.WARNING(f'Pendientes guardados en: {archivo_pendientes}'))

        self.stdout.write(self.style.SUCCESS(
            f'\nRESUMEN: {n_importados} importados · {n_saltados} ya existentes · {n_no_encontrados} no encontrados'
        ))