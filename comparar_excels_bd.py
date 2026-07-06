# comparar_excels_bd.py

import os
import django
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from gestion.models import Codigos

def main():
    archivo = input("Ruta del Excel: ").strip()
    archivo_path = Path(archivo)
    
    if not archivo_path.exists():
        print("Archivo no encontrado")
        return

    # Leer Excel
    wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
    ws = wb.active
    
    # Plan SIGMA de BD
    plan_sigma_bd = set(Codigos.objects.values_list('plan_sigma', flat=True))
    print(f"Códigos en BD: {len(plan_sigma_bd)}")
    
    # Encabezado
    encabezado = [cell.value for cell in ws[1]]
    
    # Filas nuevas (no en BD)
    filas_nuevas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        plan_sigma = fila[0]
        if plan_sigma and str(plan_sigma).strip() not in plan_sigma_bd:
            filas_nuevas.append(fila)

    # Crear nuevo Excel
    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    
    for col_num, valor in enumerate(encabezado, 1):
        ws_nuevo.cell(row=1, column=col_num, value=valor)
    
    for row_num, fila in enumerate(filas_nuevas, 2):
        for col_num, valor in enumerate(fila, 1):
            ws_nuevo.cell(row=row_num, column=col_num, value=valor)

    # Guardar
    archivo_nuevo = Path(f'{archivo_path.stem}_faltantes.xlsx')
    wb_nuevo.save(str(archivo_nuevo))

    print(f'De {ws.max_row - 1} filas, {len(filas_nuevas)} faltan en BD.')
    print(f'Archivo: {archivo_nuevo}')

if __name__ == '__main__':
    main()